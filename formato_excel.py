# -*- coding: utf-8 -*-
"""
Paquete de funciones para trabajar con exceles ya existentes. Se basa en openpyxl para trabajar con contenido y formatos.

Funciones:
    - autoajustar:               función que autoajusta el ancho de las columans de una hoja a su contenido.
    - colorear_fila:             función que colorea una fila completa de una hoja.
    - crearAutofiltro:           función que crea un autofiltro a partir de una fila dada.
    - escribirRango:             función que escribe un dato aislado o un dataframe en un rango específico de una hoja
    - formatear_hoja:            función que aplica autoajuste de ancho, colorea la cabeecera de azul corporativo y aplica autofiltro a una hoja.
    - formatear_libro:           (no operativa) cuando esté lista, funcionará como formatear_hoja, pero aplicada a todas las hojas de un libro.
    - formatear_rango:           función que formatea un rango específico. Permite especificar el formato numérico, de fuente, de borde, alineación del texto y relleno de celda.
    - guardar_fich_excel:        función que llama a la función save de openpyxl varias veces. Útil para evitar fallos si tenemos abierto el libro a guardar y se nos olvida cerrarlo.
    - incluirFormatoCondicional: función para añadir una nueva regla de formato condicional a un rango de celdas.
    - exportar_a_pdf:            (experimental) función que exporta un libro excel completo (con todas sus hojas) a pdf. Requere tener instalado el paquete pdfrw
    - limpiarRango:              función que limpia los valores (no los formatos) de una sección de filas y columnas de una hoja excel.
    
    - calcular_num_columna       función que traduce una columna en letras a índice que ocupa.
    - leer_hoja_excel_grande     función que eficienta la lectura de excel grandes mediante conversión previa a csv
    - proteger_libro             función que establece una contraseña de apertura para un archivo xlsx
"""

__version__ = '20241213.0'
__changelog__ = '''
[20241213] - Se incluye la función proteger_libro
'''

#%%
from tkinter import messagebox

import math
import openpyxl
import xlsxwriter
import sys

from openpyxl            import load_workbook, Workbook
from openpyxl.styles     import Color, PatternFill, Font, Border, Alignment, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.cell       import Cell
from openpyxl.formatting import Rule

# Set header formats:
headerFill1 = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('004481'))
headerFill2 = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('2dcccd'))
headerFill3 = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('666666'))
headerFill4 = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('B92A45'))
headerFill5 = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('9C6C01'))

redErrorFill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('FF6666'))

pijamaFill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('D4EDFC'))

FORMATOS_BASE = {
        'euro_2_dec'            : '#,##0.00 €',
        'euro_no_dec'           : '#,##0 €',
        
        'fecha_estandar'        : 'dd/mm/yyyy',
        'fecha_ordenada'        : 'yyyy-mm-dd',
        'fechahora_estandar'    : 'dd/mm/yyyy HH:MM:SS',
        'fechahora_ordenada'    : 'yyyy-mm-dd HH:MM:SS',

        'num_sep_mil_2dec'      : '#,##0.00_);[Red]-#,##0.00',
        'num_no_sep_mil_2dec'   : '###0.00_);[Red]-###0.00',
        
        'num_sep_mil_no_dec'    : '#,##0_);[Red]-#,##0',
        'num_no_sep_mil_no_dec' : '###0_);[Red]-###0',
        
        'pct_1_dec'             : '0.0%',
        'pct_2_dec'             : '0.00%',
        
        'texto'                 : '@',
        'texto_4_posic'         : '0000_@',
        'texto_14_posic'        : '00000000000000_@',
        'texto_26_posic'        : '00000000000000000000000000_@'
}

headerFont = Font(bold=True, color='FFFFFF')

noBorder = openpyxl.styles.borders.Border(
    left=openpyxl.styles.Side(border_style=None), 
    right=openpyxl.styles.Side(border_style=None), 
    top=openpyxl.styles.Side(border_style=None), 
    bottom=openpyxl.styles.Side(border_style=None),
)

###########################################################################
#%%
def autoajustar(hoja, margen=1):
    """Función que autoajusta el ancho de las columnas de una hoja
    Parámetros:
    hoja (openpyxl.Worksheet) 
        Hoja del libro abierta con openpyxl
    """
    for icol, column_cells in enumerate(hoja.columns):
        #si no hay ningún valor, salto esta columna.
        if len(column_cells)<=1:
            continue
        #si hay valores, auto-ajustamos a la longitud máxima +2 
        length = max(len(str(cell.value)) for cell in column_cells if not any(cell.coordinate in combinada for combinada in hoja.merged_cells.ranges))
        if length<1:
            length=1

        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
        f = 0
        while type(column_cells[f]) != openpyxl.cell.cell.Cell:
            f+=1
        if openpyxl.__version__ >= '2.6':
            hoja.column_dimensions[column_cells[f].column_letter].width = length+margen
        else:
            hoja.column_dimensions[column_cells[f].column].width = length+margen
   
    return hoja
###########################################################################
#%%
def colorear_fila(worksheet, num_fila, fill):
    """
    Función que colorea una fila de un worksheet
    -------
    Parametros:
        worksheet (openpyxl.Worksheet)
            hoja a la que pertenece la fila a colorear
        num_fila (int)  
            Posición de la fila en la hoja
        fill (openpyxl.styles.fills.PatternFill) 
            Objeto con el contenido del color que se le quiere dar
    """
    #Primero le damos a la cabecera su formato:
    for y in range(1, worksheet.max_column+1):
        worksheet.cell(row=num_fila, column=y).fill = fill
    return worksheet
    
#%%
def crearAutoFiltro(worksheet, num_fila = 1):
    """
    Función que añade un autofiltro en una fila determinada
    -------
    Parametros:
        worksheet (openpyxl.Worksheet)
            hoja a la que pertenece la fila a colorear
        num_fila (int) default 1
            Fila en la que poner el autofiltro
    Devuelve objeto worksheet
    """
    zona_datos  = worksheet.calculate_dimension()
    zona_filtro = 'A' + str(num_fila)
    if len(zona_datos.split(':'))>1:
        zona_filtro = zona_filtro + ':' + zona_datos.split(':')[1]
    worksheet.auto_filter.ref = zona_filtro
    
    return worksheet

    
###########################################################################
#%%
def escribirRango(libro, hoja, datos, celda='A1', header=True, index=False, autoajuste=False, formateo=False):
    """
    Función que escribe un dato o serie de datos en una excel.

    Parameters
    ----------
    libro : (str) o (openpyxl.workbook.workbook.Workbook)
        Si es (str), ruta del libro que se desea escribir. Devolverá el objeto openpyxl guardado.
        Si es un workbook, devolverá el workbook SIN GUARDAR
    hoja : (str) o (openpyxl.worksheet.worksheet.Worksheet)
        Nombre de hoja u objeto hoja en el que se desea escribir. Si la hoja no existe, se creará. Si libro es str, hoja también debe serlo.
    datos : (str), (int), (float), (bool), (pandas.DataFrame)
        Dato que escribir. Si es un dataframe, se escribirá a partir de la celda indicada en el parámetro celda
    celda : (str), optional
        Celda en la que escribir los datos, en formato hoja de cálculo. The default is 'A1'.
    header : (bool), optional
        Indica si debe escribir la cabecera del dataframe o no. The default is True.
    index : (bool), optional
        Indica si debe escribir el índice del dataframe o no. The default is False.
    autoajuste : (bool), optional
        Indica si debe autoajustar los anchos de la hoja tras la escritura. The default is False.
    formateo : (bool), optional
        Indica si debe aplicar formato corporativo automático a la hoja tras escribirla. The default is False.

    Returns
    -------
    None. La función guardará automáticamente el libro

    """
    import openpyxl
    
    #Comprobación de validez de los parámetros
    if not libro:
        raise Exception('El parámetro libro no es válido. Debe ser un nombre de archivo válido o un objeto openpyxl.workbook.workbook.Workbook')
    
    fichero = None
    if type(libro) == str:
        fichero = libro
        libro = openpyxl.load_workbook(fichero)
    elif type(libro) != openpyxl.workbook.workbook.Workbook:
        raise Exception('El parámetro libro no es válido. Debe ser un nombre de archivo válido o un objeto openpyxl.workbook.workbook.Workbook')
    
    #Hoja:
    if not hoja:
        raise Exception('El parámetro hoja no es válido. Debe ser un nombre de hoja válida o un objeto openpyxl.worksheet.worksheet.Worksheet')
    if type(hoja) == str:
        if hoja not in libro.sheetnames:
            hoja = libro.create_sheet(hoja)
        else:
            hoja = libro[hoja]
    elif type(hoja) == openpyxl.worksheet.worksheet.Worksheet:
        if fichero:
            raise Exception('Si se proporciona un nombre de fichero, la hoja debe ser también str, no un objeto openpyxl.')
    else:
        raise Exception('El parámetro libro no es válido. Debe ser un nombre de archivo válido o un objeto openpyxl.worksheet.worksheet.Worksheet')
    
    #Datos:
    import pandas
    pivote = None
    if type(datos) != pandas.core.frame.DataFrame:
        #Si no es un dataframe, escribimos el dato directamente a la celda indicada
        hoja[celda] = datos
    else:
        #Es un dataframe. Hay que escribirlo A PARTIR de dicha celda.
        pivote = openpyxl.utils.coordinate_to_tuple(celda)

        from openpyxl.utils.dataframe import dataframe_to_rows
        filas = dataframe_to_rows(datos, index=index, header=header)
        
        for ifila, fila in enumerate(filas, pivote[0]):
            for icol, dato in enumerate(fila, pivote[1]):
                 hoja.cell(row=ifila, column=icol, value=dato)
    
    if autoajuste:
        autoajustar(hoja)
    if formateo:
        if pivote:
            formatear_hoja(hoja, num_fila_cabecera=pivote[0])
        else:
            formatear_hoja(hoja)
    
    if fichero:
        libro.save(fichero)
    return libro


###########################################################################
#%%
def formatear_hoja(worksheet, num_fila_cabecera = 1):
    """
    Función que formatea una hoja: 
    Ajusta el tamaño de las columnas al contenido de las celdas
    Ajusta el tamaño de la primera fila (cabecera) al contenido de la misma
    ------
    Parámetros:
        worksheet (openpyxl.worksheet.worksheet.Worksheet) 
            Hoja a la que queremos dar formato
    """

    #Primero le damos a la cabecera su formato:
    if type(num_fila_cabecera) == int:
        filas_cabecera = [num_fila_cabecera]
    elif type(num_fila_cabecera) == list:
        filas_cabecera = num_fila_cabecera.copy()
    else:
        return None
    max_fila_cabecera = max(filas_cabecera)
        
    for fila in filas_cabecera:
        for y in range(1, worksheet.max_column+1):
            worksheet.cell(row=fila, column=y).fill = headerFill1
            worksheet.cell(row=fila, column=y).font = headerFont
            worksheet.cell(row=fila, column=y).border = noBorder            
    
    #Luego ajustamos el ancho de las columnas al contenido (sin contar la cabecera)
    anchos = {}
    for icol, column_cells in enumerate(worksheet.columns):
        #si no hay ningún valor, salto esta columna.
        if len(column_cells)<1:
            continue
        
        #Si solo hay un valor (el nombre de columna), autoajustamos al mismo:
        length = 0
        if len(column_cells)==1:
            length = max([0] + [len(str(cell.value)) for cell in column_cells if not any(cell.coordinate in combinada for combinada in worksheet.merged_cells.ranges)])
        
        else:
            #si hay más valores, auto-ajustamos a la longitud máxima +2. Si hay más de 1.000 filas sólo miro las 1.000 primeras
            ultimo_elto = min([len(column_cells)+1, 1001])
            length = max([0]+[len(str(cell.value)) for cell in column_cells[max_fila_cabecera:ultimo_elto] if not any(cell.coordinate in combinada for combinada in worksheet.merged_cells.ranges)])
        
        if length<4:
            length=4
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
        f = 0
        while type(column_cells[f]) != openpyxl.cell.cell.Cell:
            f+=1
        if openpyxl.__version__ >= '2.6':
            worksheet.column_dimensions[column_cells[f].column_letter].width = length+2
        else:
            worksheet.column_dimensions[column_cells[f].column].width = length+2
        anchos[icol] = length+2

    
    #Filas de cabeceras: wrap_text y altura:
    altura_fila_texto= 15 #esto es el default

    for fila in filas_cabecera:
        fila_cabecera=worksheet[fila]
        
        #Para empezar la cabecera tiene altura para que quepa 1 línea de texto. En 
        #función del ancho de la columna vamos a calcular cuántas líneas harían falta
        #para ver la cabecera entera. Nos quedaremos con el máximo de todas las columnas
        
        filas_texto_cab=1
        for icol, cell in enumerate(fila_cabecera) :
            cell.alignment = Alignment(wrap_text=True, vertical = "center")
            try:
                longitud_celda = len (cell.value)
            except:
                longitud_celda = 1

            
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
            # if openpyxl.__version__ >= '2.6':
            #     ancho_columna = worksheet.column_dimensions[cell.column_letter].width
            # else:
            #     ancho_columna = worksheet.column_dimensions[cell.column].width
            if (icol not in anchos) or (anchos[icol] is None) or (anchos[icol] == 0):
                continue
                
            ancho_columna = anchos[icol]
            if ancho_columna == None or ancho_columna == 0:
                continue                                               
            
            filas_texto_necesarias = math.ceil(longitud_celda / ancho_columna)
            if filas_texto_necesarias>filas_texto_cab:
                filas_texto_cab=filas_texto_necesarias
        
        #Tampoco queremos una cabecera con 40 líneas de texto, por eso como mucho dejamos 6
        worksheet.row_dimensions[fila].height = altura_fila_texto* min(filas_texto_cab,6)
    
    #Inmovilizamos la cabecera
    worksheet.freeze_panes = 'A'+str(max_fila_cabecera+1)
    
    #añadir auto filtro
    zona_datos=worksheet.calculate_dimension()
    zona_filtro = 'A'+str(max_fila_cabecera)
    if len(zona_datos.split(':'))>1:
        zona_filtro=zona_filtro +':'+ zona_datos.split(':')[1]
    worksheet.auto_filter.ref = zona_filtro
    
    return worksheet

###########################################################################
#%%
def formatear_libro(libro, hoja = None, save=True, close=False, num_fila_cabecera = 1, **kwargs):
    """
    Función que formatea un libro de excel conforme a las directrices dadas

    Parameters
    ----------
    libro : (str) o (openpyxl.Workbook)
        Libro que se desea formatear. Se puede pasar tanto la ruta al mismo como el objeto openpyxl si ya lo tenemos.
    hoja : None, (str), (list) o (openpyxl.worksheet.worksheet.Worksheet), optional
        Hoja que se desea formatear. Si es None, formateará todas las hojas.
        Si es (str) será el nombre de la hoja que se desea formatear.
        Si es (openpyxl.worksheet.worksheet.Worksheet) será la hoja que se desea formatear.
        Si es (list) tendrá las hojas (nombres u objetos openpyxl) que se desean formatear.
    save : bool (default True), optional
        Indica si debe guardar el libro antes de devolver el resultado.
    close: bool (default False), optional
        Indica si debe cerrar el libro antes de devolver el resultado.
    num_fila_cabecera: int (default 1), optional
        Indica en qué fila están las cabeceras de la tabla
    **kwargs : Resto de parámetros admitidos en la función formatear_rango. 
        Si no se pasa ninguno, se llamará a la función formatear_hoja para dar un formato estándar.

    Returns
    -------
    openpyxl.Workbook.

    """
    import openpyxl 
    
    if type(libro) == str:
        #El libro es una ruta.
        libro = openpyxl.load_workbook(libro)
    pass

###########################################################################
#%%
def formatear_rango(worksheet, rango, formato='num_sep_mil_2dec', alineacion=None, fuente=None, borde=None, relleno=None):
    """
    Función que da formato a la celda o rango de celdas pasadas
    --------
    Parámetros:
        worksheet (openpyxl.worksheet.worksheet.Worksheet)
            Hoja a la que pertenecen las celdas que se quieren formatear
        rango (str)
            Celdas a la que se quiere dar formato
        formato (str)
            Formato del texto, por defecto = num_sep_mil_2dec
        alineacion ( openpyxl.styles.alignment.Alignment ) 
            Objeto con la alineación que se le quiere dar a las celdas
        fuente ( openpyxl.styles.fonts.Font) 
            Objeto con el formato de la fuente
        Borde ( openpyxl.styles.borders.Border) 
            Objeto con el tamaño del borde las celdas
        relleno ( openpyxl.styles.fills.PatternFill) 
            Objeto con el contenido del relleno de las celdas       
    """
    
    import openpyxl
    
    formatos_base = {
            'euro_2_dec'            : '#,##0.00 €',
            'euro_no_dec'           : '#,##0 €',
            
            'fecha_estandar'        : 'dd/mm/yyyy',
            'fecha_ordenada'        : 'yyyy-mm-dd',
            'fechahora_estandar'    : 'dd/mm/yyyy HH:MM:SS',
            'fechahora_ordenada'    : 'yyyy-mm-dd HH:MM:SS',

            'num_sep_mil_2dec'      : '#,##0.00_);[Red]-#,##0.00',
            'num_no_sep_mil_2dec'   : '###0.00_);[Red]-###0.00',
            
            'num_sep_mil_no_dec'    : '#,##0_);[Red]-#,##0',
            'num_no_sep_mil_no_dec' : '###0_);[Red]-###0',
            
            'pct_1_dec'             : '0.0%',
            'pct_2_dec'             : '0.00%',
            
            'texto'                 : '@',
            'texto_4_posic'         : '0000_@',
            'texto_14_posic'        : '00000000000000_@',
            'texto_26_posic'        : '00000000000000000000000000_@'
    }
    
    if formato in formatos_base.keys():
        formato = formatos_base[formato]
    
    def aplicarformato(celda, formato, alineacion, fuente, borde, relleno):
        celda.number_format = formato
        
        if alineacion is not None:
            celda.alignment = alineacion
            
        if fuente is not None:
            celda.font = fuente
            
        if borde is not None:
            celda.border= borde
            
        if relleno is not None:
            celda.fill = relleno 
        return celda
        
    try:
        tipos_celda = (openpyxl.cell.cell.Cell, openpyxl.cell.cell.MergedCell)
        rango_celdas = worksheet[rango]
        
        #Si sólo es 1 celda le aplico los formatos directamente
        if type(rango_celdas) in tipos_celda:
            celda=rango_celdas
            
            celda = aplicarformato(celda, formato, alineacion, fuente, borde, relleno)
            
            return worksheet
        else:
            #Si es un rango, tengo que recorrerlo celda por celda
            for grp_celdas in worksheet[rango]:
                if type(grp_celdas) in tipos_celda:
                    #Es una sola celda, se le aplica el formato
                    celda = aplicarformato(grp_celdas, formato, alineacion, fuente, borde, relleno)
                else:
                    #Es un grupo de celdas, hay que recorrer cada fila
                    for celda in grp_celdas:
                        celda = aplicarformato(celda, formato, alineacion, fuente, borde, relleno)

    except Exception as ex:
        print(ex)
    return worksheet

###########################################################################
#%%
def guardar_fich_excel (excel_book, num_intentos, nombre_fichero, ventana_ppal):
    """
    Función para guardar un libro en un fichero excel 
    ---------
    Parámetros:
        excel_book (str) o (openpyxl.Workbook)
        Libro que se desea formatear. Se puede pasar tanto la ruta al mismo como el objeto openpyxl si ya lo tenemos.
    num_intentos (int)
        Numero de intentos para guadar el archivo
    nombre_fichero (str)
        Ruta que se le quiere dar al nuevo fichero generado
    ventana_ppal (TODO)
        TODO
    """
    retry="yes"
    i=0
    
    while retry=="yes" and i<3:
        try:
            excel_book.save(nombre_fichero)
        except Exception as e:
            #Si ha fallado preguntamos si quiere volver a intentarlo. 
            #   Si dice que no->salir con error.
            #   Si dice que sí -> incrementamos el contador y volvemos a intentarlo
            print(e)
            retry = messagebox.askquestion ('Error al Guardar',
                                               'No se puede guardar el fichero "'+nombre_fichero+'"\n' 
                                               + '¿Quieres volver a intentar guardarlo?',
                                               icon = 'error')
            if retry == 'yes':
               i+=1
            else:
                return False
        else:
            #Si no ha generado excepción es que ha ido todo bien. Salimos con OK
            return True
    
    #Si llega a este punto es que ha superado el número de intentos máximos -> salir con error
    messagebox.showerror("Error al Guardar", 'Se ha superado el límite de intentos de guardar el fichero')
    return False

###########################################################################
#%%
def incluirFormatoCondicional(worksheet, rango, formula_celda1, color_fondo, color_texto='000000', negrita = False):
    """
    Función para añadir una nueva regla de formato condicional a un rango de celdas. La regla se añade, es decir, si ya
    había reglas de formato condicional en esa celda no se borran, sólo se añade una nueva. El orden en que se añaden es
    relevante, ya que en cuanto cumpla una regla, ese será el formato que aplique.

    Parameters
    ----------
    worksheet (openpyxl.worksheet.worksheet.Worksheet)
        Hoja a la que pertenecen las celdas que se quieren formatear
    rango (str)
        Rango de celdas a la que se quiere dar formato en notación excel. Ejemplos: 'A2', 'B3:W45', 'C:D'
    formula_celda1 : (str)
        Fórmula a aplicar en la 1a celda. Si se quiere fijar algún valor es importante el uso de '$'.
        Ejemplos: Supongo que mi rango es 'A2:Z100' 
            -Si quiero que toda la fila se coloree cuando en la columna A tenga el valor "Colorear"
             la fórmula será: '$A2="Colorear"' (con $ en la A porque quiero fijar la columna)
            -Si quiero que se coloreen las celdas con valor > 100 la fórmula será 'A2>100' (sin $)
    color_fondo : (str)
        Código de color RGB sin '#'. Ejemplos: Blanco = 'FFFFFF', Rojo = 'FF0000'.
    color_texto : (str), optional
        Código de color RGB sin '#'. The default is '000000' (negro).
    negrita : (bool), optional
        Indicar si queremos que el texto se resalte en negrita. The default is False.

    Returns
    -------
    None.

    """
    
    
    relleno = PatternFill( bgColor=Color(color_fondo)) 
    fuente = Font(bold=negrita, color=color_texto)
    formato_condicional = DifferentialStyle(fill=relleno, font=fuente)
    regla = Rule(type="expression", dxf=formato_condicional, stopIfTrue=True)
    regla.formula = [formula_celda1]
    worksheet.conditional_formatting.add(rango, regla)
    
    pass
    
###########################################################################

#%%
def limpiarRango(hoja, pivote='A1', numFilas=None, numColumnas=None):
    """
    Función que limpia los datos de un rango de una hoja excel

    Parameters
    ----------
    hoja : openpyxl.worksheet.worksheet.Worksheet
        Hoja en la que se encuentra el rango que limpiar.
    pivote : str, optional
        Esquina superior izquierda del rango, en formato coordenada excel. The default is 'A1'.
    numFilas : int, optional
        Número de filas que limpiar. None si hay que limpiar todas. The default is None.
    numColumnas : int, optional
        Número de columnas que limpiar None si hay que limpiar todas. The default is None.

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet. La hoja modificada

    """
    import openpyxl
    if type(hoja) != openpyxl.worksheet.worksheet.Worksheet:
        raise Exception('El parámetro hoja debe ser una openpyxl.worksheet.worksheet.Worksheet')
    
    tpivote = openpyxl.utils.coordinate_to_tuple(pivote)

    try:
        if numFilas is None:
            numFilas = hoja.max_row + 1
        elif int(numFilas) > 0:
            numFilas = tpivote[0] + int(numFilas)
    except:
        raise Exception('numFilas debe ser un número entero positivo')

    try:
        if numColumnas is None:
            numColumnas = hoja.max_column + 1
        elif int(numColumnas) > 0:
            numColumnas = tpivote[1] + int(numColumnas)
    except:
            raise Exception('numColumnas debe ser un número entero positivo')
    
    #Recorremos todas las celdas que hay que limpiar
    for nFila in range(tpivote[0], numFilas):
        for nCol in range(tpivote[1], numColumnas):
            hoja.cell(row=nFila, column=nCol).value = None
    
    return hoja
    




##############################################################################
#%%



############        LECTURA DE EXCELS GRANDES          #######################

def calcular_num_columna(col_en_letras, base0=True):
    '''
    Función que en función de una columna en letras devuelve el número de columna. 
    Puede devolver el número de columna empezando a contar por 0 (base0=True) o por 1 (base0=False).
    Ejemplo: 
        * calcular_num_columna('A', base0=True)  --> 0
        * calcular_num_columna('A', base0=False) --> 1
        * calcular_num_columna('AZ', base0=True)  --> 51
        * calcular_num_columna('AZ', base0=False) --> 52

    Parameters
    ----------
    col_en_letras : str
        Letra de la columna a convertir.
    base0 : bool, optional
        Indicar si se quiere que la columna A se considere la 0 (True) o la 1 (False). The default is True.

    Returns
    -------
    int
        Número de columna de las letras enviadas.

    '''
    num_col = 0
    col_en_letras = col_en_letras.upper()
    
    for indice, letra in enumerate(col_en_letras[::-1]):
        num_col += (ord(letra)-ord('A')+1) * ((ord('Z')-ord('A')+1)**indice)
        
    if base0:
        return num_col-1
    
    return num_col
    

def calcular_letra_columna(col_en_numero):
    '''
    Función que recibe un número de columna en base 1 y devuelve la(s) letra(s) correspondiente en hoja de cálculo.
    Ejemplo: 
        * calcular_letra_columna(1)  --> 'A'
        * calcular_letra_columna(10) --> 'J'
        * calcular_letra_columna(26) --> 'Z'
        * calcular_letra_columna(30) --> 'AD'
        * calcular_letra_columna(81) --> 'CC'
        * calcular_letra_columna(0)  --> None
        * calcular_letra_columna(-5) --> None

    Parameters
    ----------
    col_en_numero : int
        Número de columna a traducir a letra en base 1

    Returns
    -------
    str con la letra de la columna en hoja de cálculo.
    None si el parámetro es <=0.

    '''
    if col_en_numero<=0:
        return None
    
    letra_col = ''
    while col_en_numero > 0:
        col_en_numero, resto = divmod(col_en_numero - 1, 26)
        letra_col = chr(ord('A') + resto) + letra_col
        
    return letra_col



def leer_hoja_excel_grande(nom_fichero, hoja=1, cols_numericas = None, cols_fechas=None, filas_cabecera=0, tipos_dato=None):
    '''
    Función que lee una pestaña de un excel grande convirtiéndolo primero a CSV.
    Necesita que EXCEL esté instalado (no sirve libre office ni open office).
    Interpreta las columnas numéricas y de fecha. 

    Parameters
    ----------
    nom_fichero : str
        Ruta completa del fichero a leer
    hoja : str o int, optional
        Hoja a leer. Puede ser el nombre o posición (1-based). The default is 1.
    cols_numericas : list, optional
        Lista de columnas que deben generarse como número en formato hoja de cálculo. 
        Ejemplo: ['A', 'B', 'C', 'W']
        The default is None.
    cols_fechas : list, optional
        Lista de columnas que deben generarse como fecha en formato hoja de cálculo. 
        Ejemplo: ['D', 'AB']
        The default is None.
    filas_cabecera : list o int , optional
        Filas del excel en las que aparece la cabecera. 
        Si no se indica, se toma la primera fila como cabecera
        The default is 0.
    tipos_dato : dict o tipo de dato, optional
        Diccionario con claves nombre_columna : tipo_dato para hacer la lectura. 
        The default is None.

    Returns
    -------
    pandas DataFrame con el dataframe leído.

    '''
       
    
    import os
    import pandas as pd
    #Líneas que nos dan la ruta de la carpeta de descargas del usuario
    from winreg import OpenKey, HKEY_CURRENT_USER, QueryValueEx
    with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
        ruta_descargas_usuario = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
    
    
    #Preparamos el script para pasar el excel a csv y leer más rápido
    vbscript="""
        if WScript.Arguments.Count < 3 Then
            WScript.Echo "Please specify the source and the destination files. Usage: ExcelToCsv <xls/xlsx source file> <csv destination file> <worksheet name/number> [range numeric] [range date] [range delete ordered]"
            Wscript.Quit
        End If
        
        
        On Error Resume Next
        
        Set objFSO = CreateObject("Scripting.FileSystemObject")
        
        src_file = objFSO.GetAbsolutePathName(Wscript.Arguments.Item(0))
        dest_file = objFSO.GetAbsolutePathName(WScript.Arguments.Item(1))
        worksheet_name = WScript.Arguments.Item(2)
        
        If IsNumeric(worksheet_name) Then
            worksheet_name = CLng(worksheet_name)
        End If
    	
        
        Dim oExcel
        Set oExcel = CreateObject("Excel.Application")
        
        Dim oBook
        Set oBook = oExcel.Workbooks.Open(src_file)
        oBook.Worksheets(worksheet_name).Activate
    	
    	'Si se han pedido ciertas columnas como numericas
    	if WScript.Arguments.Count > 3 Then
    		rangos = Split(WScript.Arguments.Item(3),",")
    		
    		for each rango in rangos
    			Set objRange = oExcel.Range(rango)
    			objRange.NumberFormat = "0.00"
    		next		
    	End If
    	
    	'Si se han pedido ciertas columnas como fecha
    	if WScript.Arguments.Count > 4 Then
    		rangos = Split(WScript.Arguments.Item(4),",")
    		
    		for each rango in rangos
    			Set objRange = oExcel.Range(rango)
    			objRange.NumberFormat = "dd/mm/yyyy"
    		next		
    	End If
        
    	'Si se quieren eliminar ciertas columnas
    	if WScript.Arguments.Count > 5 Then
    		rangos = Split(WScript.Arguments.Item(5),",")
    		
    		for k = Ubound(rangos) to LBound(rangos) step -1
    			rango = rangos(k)
    			Set objRange = oExcel.Range(rango)
    			objRange.EntireColumn.Delete
    		Next
    	End If
    	
        oExcel.DisplayAlerts = False
        'SaveAs fich_salida, FileFormat, Password, WriteResPassword, ReadOnlyRecommended, CreateBackup, AccessMode, ConflictResolution, AddToMru, TextCodepage, TextVisualLayout, Local
        oBook.SaveAs dest_file,       6,    Null ,             Null,                Null,        False,       Null,               Null,     Null,         Null,             Null, False
        
        oBook.Close False
        oExcel.Quit
    
        """
    f = open(ruta_descargas_usuario + r'\ExcelToCSV.vbs','w')
    f.write(vbscript)
    f.close()
    
    fich_sin_ruta = nom_fichero.replace('\\','/').split('/')[-1]
    
    cols_num = ''
    if cols_numericas is not None and len(cols_numericas)>0:
        cols_num = ','.join([col+':'+col for col in cols_numericas])
    
    cols_fec = ''  
    parse_dates = None
    dayfirst=False      
    if cols_fechas is not None and len(cols_fechas)>0:
        cols_fec = ','.join([col+':'+col for col in cols_fechas])
        parse_dates=[ calcular_num_columna(col, base0=True) for col in cols_fechas]
        dayfirst=True
        
    
    #Convertir el excel a csv para hacer una lectura más rápida
    try:
        comando = ruta_descargas_usuario + r'\ExcelToCSV.vbs "'+ nom_fichero \
            +'" "' + ruta_descargas_usuario + '\\' +fich_sin_ruta.replace('.xlsx', '.csv') \
            + '" "' + str(hoja) +'" "'+cols_num+'" "'+cols_fec+'"'
        print(comando)
        os.system(comando)
    except Exception as e:
        print('Error en leer_excel_grande:', e)
        return None
    
    
    return pd.read_csv(ruta_descargas_usuario + '\\' +fich_sin_ruta.replace('.xlsx', '.csv'),
                       sep=',', encoding='latin1', header=filas_cabecera, dtype=tipos_dato, 
                       parse_dates=parse_dates, dayfirst=dayfirst)
    
###########################################################################
#%%
def rgb_a_hex(r,g,b):
    """
    Convierte una tupla de valores enteros RGB a una cadena hexadecimal.
    """
    # Convierte cada valor de RGB en una cadena hexadecimal de 2 dígitos
    hex_code = '{:02X}{:02X}{:02X}'.format(r, g, b)
    # Devuelve la cadena hexadecimal
    return hex_code

#%% proteger_libro
def proteger_libro(ruta_libro:str, password:str, ruta_salida:str=None, confirmar_sobreescritura:bool=False) -> None:
    """
    Protege un libro de Excel con una contraseña.

    Parameters:
    ruta_libro               : (str)          - Ruta del archivo Excel que se quiere proteger.
    password                 : (str)          - Contraseña de protección.
    ruta_salida              : (str)  = None  - Ruta donde guardar el fichero. Si no se proporciona, guardará en ruta_libro
    confirmar_sobreescritura : (bool) = False - Indica si debe mostrar una alerta si existe un fichero con el mismo nombre de guardado.
    """
    import win32com.client as win32

    # Inicializamos Excel
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    if confirmar_sobreescritura:
        excel.DisplayAlerts = True
    else:
        excel.DisplayAlerts = False

    if ruta_salida is None:
        ruta_salida = ruta_libro

    try:
        # Abrimos el libro existente
        wb = excel.Workbooks.Open(ruta_libro)

        # Guardamos con la contraseña
        wb.SaveAs(
            ruta_salida,
            FileFormat=51,  # XLSX
            Password=str(password)
        )

        print(f"El libro ha sido protegido con contraseña: {ruta_salida}")

    except Exception as e:
        print(f"Error al proteger el libro: {e}")

    finally:
        # Cerramos el libro y Excel
        wb.Close(SaveChanges=False)
        excel.Quit()

# %%
